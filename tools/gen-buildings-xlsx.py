# -*- coding: utf-8 -*-
"""生成建筑可配置表格 data/buildings-config.xlsx(按类别分 Sheet)
数据来源:data/人工核查表.xlsx(用户 2026-08 录入/重写,原版 Anno 1800 数值,权威源)
- 从 住宅类/生产类/服务类 子表读取,按映射转换(中文名→引擎 id)
- 备注含"暂不实装"的建筑跳过
用法:python tools/gen-buildings-xlsx.py
修改表格后运行:python tools/gen-buildings-js.py 重新生成引擎数据
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = "data/人工核查表.xlsx"
DST = "data/buildings-config.xlsx"

# 中文名 → 引擎 id(生产/服务)
BID = {
    "原木厂": "sawmill", "木板厂": "boardmill", "渔场": "fishery", "土豆农场": "potatoField",
    "烈酒厂": "distillery", "绵羊牧场": "sheepFarm", "纺织厂": "tailor", "陶土矿场": "clayPit",
    "砖厂": "brickworks", "猪牧场": "pigFarm", "香肠生产厂": "sausageFactory", "谷物农场": "grainFarm",
    "磨坊": "mill", "面包店": "bakery", "炭窑": "charcoalKiln", "铁矿": "ironMine",
    "高炉": "blastFurnace", "钢材厂": "steelWorks", "精炼厂": "renderingWorks", "肥皂厂": "soapFactory",
    "啤酒花农场": "hopFarm", "麦芽加工厂": "maltWorks", "酿酒厂": "brewery", "砂石采集场": "sandPit",
    "玻璃厂": "glassworks", "窗户厂": "windowFactory", "牛牧场": "cattleFarm", "红椒农场": "pepperFarm",
    "工匠厨房": "artisanKitchen", "罐头工厂": "cannery", "煤矿": "coalMine", "缝纫机厂": "sewingMachineFactory",
    "混凝土厂": "concreteWorks", "蒸汽机厂": "steamWorks", "硝石采集场": "saltpeterWorks",
    "葡萄农场": "grapeFarm", "石灰岩矿": "limestoneMine", "锌矿": "zincMine", "铜矿": "copperMine",
    "金矿": "goldMine", "黄铜冶炼厂": "brassWorks", "香槟酒厂": "champagneCellar", "眼镜厂": "glassesWorks",
    "怀表工厂": "watchFactory", "灯丝厂": "filamentWorks", "灯泡工厂": "bulbFactory",
    "薄木片工厂": "veneerWorks", "留声机工厂": "phonographWorks", "炼油厂": "oilRefinery",
    "仓库": "warehouse", "市场": "market", "酒吧": "bar", "学校": "school", "教堂": "church",
    "大学": "university", "剧院": "theater", "银行": "bank", "燃油发电厂": "powerPlant",
    "会员俱乐部": "club",
}
# 住宅:英文 ID → 引擎 id
RESID = {
    "Farmer Residence": "residence", "Worker Residence": "residenceWorkers",
    "Artisan Residence": "residenceArtisans", "Engineer Residence": "residenceEngineers",
    "Investor Residence": "residenceInvestors",
}
# 商品中文名 → id
GOOD = {
    "原木": "log", "木材": "wood", "鱼": "fish", "土豆": "potato", "烈酒": "schnapps",
    "羊毛": "wool", "工作服": "workclothes", "陶土": "clay", "砖块": "brick", "猪": "pig",
    "香肠": "sausage", "谷物": "grain", "小麦粉": "flour", "面包": "bread", "煤": "coal",
    "铁矿石": "ironOre", "钢铁": "steelBar", "钢材": "steel", "动物性油脂": "lard", "肥皂": "soap",
    "啤酒花": "hops", "麦芽": "malt", "啤酒": "beer", "石英砂": "sand", "玻璃": "glass",
    "窗户": "windows", "牛肉": "beef", "红椒": "pepper", "红椒炖肉": "cannedFood", "罐头": "canned",
    "缝纫机": "sewingMachine", "钢筋混凝土": "concrete", "蒸汽机": "steamEngine", "硝石": "saltpeter",
    "葡萄": "grapes", "水泥": "cement", "锌矿石": "zincOre", "铜矿石": "copperOre",
    "金矿石": "goldOre", "黄金": "goldOre", "黄铜": "brass", "香槟": "champagne", "眼镜": "glasses",
    "灯丝": "filament", "灯泡": "lightBulb", "薄木片": "veneer", "留声机": "phonograph",
    "石油": "oil", "怀表": "pocketWatch",
}
# 服务类型
SVC = {
    "仓库": "warehouse", "市场": "market", "酒吧": "bar", "学校": "school", "教堂": "church",
    "大学": "university", "剧院": "theater", "银行": "bank", "燃油发电厂": "electricity",
    "会员俱乐部": "club",
}
# 特殊地形
TERRAIN = {
    "海岸": "coast", "陶土矿": "clay", "铁矿": "iron", "铜矿": "copper", "金矿": "gold",
    "煤矿": "coal", "锌矿": "zinc", "石灰岩矿": "limestone",  # [顺序8] 新矿脉
}

wb = openpyxl.load_workbook(SRC)
out = openpyxl.Workbook()

# ---------- 住宅类 ----------
ws = wb["住宅类"]
o = out.active
o.title = "住宅类"
headers = ["建筑ID", "名称", "所属阶层", "宽", "高", "成本-金币", "成本-木材", "住宅容量",
           "升级目标建筑ID", "升级消耗-木材", "升级消耗-砖块", "升级消耗-钢材", "升级消耗-窗户", "升级消耗-混凝土", "备注"]
o.append(headers)
res_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[1]:
        continue
    rid = RESID.get(str(row[0] or ""))
    if not rid:
        print("!! 未映射住宅:", row[0]); continue
    res_rows.append([
        rid, row[1], row[2], row[3], row[4], row[5] or 0, row[6] or 0, row[7] or 0,
        RESID.get(str(row[8] or ""), "") or "", row[9] or 0, row[10] or 0, row[11] or 0, row[12] or 0, row[13] or 0, "",
    ])
for r in res_rows:
    o.append(r)

# ---------- 生产类 ----------
ws = wb["生产类"]
o = out.create_sheet("生产类")
headers = ["建筑ID", "名称", "所属阶层", "宽", "高", "成本-金币", "成本-木材", "成本-砖块", "成本-钢材",
           "成本-窗户", "成本-混凝土", "维护费/分钟", "劳动力需求", "产出-商品", "生产周期(秒)", "产出-数量",
           "消耗-商品", "消耗-数量", "特殊地形", "未开发区域半径", "服务类型", "服务半径(路距离格)", "备注"]
o.append(headers)
TIER_COL = {"farmers": 12, "workers": 13, "artisans": 14, "engineers": 15}  # 劳动力列(1-based)
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[1]:
        continue
    note = str(row[25] or "")
    if "暂不实装" in note:
        print("跳过(暂不实装):", row[1]); continue
    bid = BID.get(str(row[1]))
    if not bid:
        print("!! 未映射生产建筑:", row[1]); continue
    tier = row[2]
    # 用户表列(0-index):16 产出-商品 / 17 周期 / 18 数量 / 19 消耗-商品 / 20 消耗-数量 / 21 特殊地形 / 24 未开发半径 / 25 备注
    out_g = GOOD.get(str(row[16] or "").strip(), "")
    inp_raw = str(row[19] or "").strip()
    inp_g = ",".join(GOOD.get(g.strip(), g.strip()) for g in inp_raw.split(",") if g.strip()) if inp_raw else ""
    inp_q = row[20] if row[20] is not None else 0
    radius = row[24] if row[24] is not None else 0
    labors = [row[12], row[13], row[14], row[15]]  # 用户表 1-based 13-16 = 劳动力4列(0-index 12-15)
    wf = 0
    if tier in TIER_COL:
        wf = labors[TIER_COL[tier] - 12] or 0
    note = str(row[25] or "")
    o.append([
        bid, row[1], tier, row[3], row[4], row[5] or 0, row[6] or 0, row[7] or 0, row[8] or 0,
        row[9] or 0, row[10] or 0, row[11] or 0, wf, out_g, row[17] or 0, row[18] or 1,
        inp_g, inp_q, TERRAIN.get(str(row[21] or "").strip(), "") if row[21] else "", radius, "", 0, note,
    ])

# ---------- 服务类 ----------
ws = wb["服务类"]
o = out.create_sheet("服务类")
headers = ["建筑ID", "名称", "所属阶层", "宽", "高", "成本-金币", "成本-木材", "成本-砖块", "成本-钢材",
           "成本-窗户", "成本-混凝土", "维护费/分钟", "服务类型", "服务半径(路距离格)", "维护所需石油/分钟", "备注"]
o.append(headers)
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[1]:
        continue
    note = str(row[15] or "") if len(row) > 15 else ""
    if "暂不实装" in note:
        print("跳过(暂不实装):", row[1]); continue
    bid = BID.get(str(row[1]))
    if not bid:
        print("!! 未映射服务建筑:", row[1]); continue
    svc = SVC.get(str(row[1]), "")
    o.append([
        bid, row[1], row[2], row[3], row[4], row[5] or 0, row[6] or 0, row[7] or 0, row[8] or 0,
        row[9] or 0, row[10] or 0, row[11] or 0, svc, row[14] or 0, row[12] or 0, note,
    ])

# ---------- 使用说明 ----------
o = out.create_sheet("使用说明")
for line in [
    "蒸汽都市 建筑参数可配置表(由 data/人工核查表.xlsx 生成,勿手改本表)",
    "",
    "1. 修改数值:编辑 data/人工核查表.xlsx(权威源),然后运行:",
    "   python tools/gen-buildings-xlsx.py   # 重新生成本表",
    "   python tools/gen-buildings-js.py     # 重新生成引擎数据",
    "2. 需求数据:人工核查表「需求」子表 → tools/gen-needs-js.py → needs-data.js",
    "3. 备注含「暂不实装」的建筑不会被导入",
    "4. 黄色列 = 可编辑(生成后);生成器每次覆盖本表",
] :
    o.append([line])

# 表头样式
for sn in ["住宅类", "生产类", "服务类"]:
    ws2 = out[sn]
    for c in ws2[1]:
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
    widths = [16, 12, 10, 6, 6, 9, 9, 9, 9, 9, 9, 11, 11, 13, 11, 9, 13, 9, 10, 12, 10, 13, 10]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

out.save(DST)
print("OK:", DST, "| 住宅", len(res_rows), "生产", ws.max_row - 1 - 0, "(跳过暂不实装)", "服务", wb['服务类'].max_row - 1)
