# -*- coding: utf-8 -*-
"""读取 data/buildings-config.xlsx(4 个分类 Sheet)→ 生成 src/engine/data/buildings-data.js
游戏内建筑参数由该 Excel 决定(人工可改)。用法:
    python tools/gen-buildings-js.py
生成后刷新游戏生效(注意 bump index.html 资源版本号 ?v=N)
"""
import json
import openpyxl

XLSX = "data/buildings-config.xlsx"
OUT = "src/engine/data/buildings-data.js"

wb = openpyxl.load_workbook(XLSX)
BUILDINGS = {}

for sheet_name in ["住宅类", "生产类", "服务类"]:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        continue
    headers = [str(h).strip() if h else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}

    def gv(r, name):
        i = idx.get(name)
        if i is None:
            return None
        return r[i]

    for r in rows[1:]:
        if not r[0]:
            continue
        bid = str(r[0]).strip()
        name = str(gv(r, "名称") or "").strip()
        tier_raw = gv(r, "所属阶层")
        tier = (str(tier_raw).strip() if tier_raw not in (None, "") else None)
        w = int(gv(r, "宽") or 0)
        h = int(gv(r, "高") or 0)
        cost = {}
        if gv(r, "成本-金币"): cost["coin"] = int(gv(r, "成本-金币"))
        if gv(r, "成本-木材"): cost["wood"] = int(gv(r, "成本-木材"))
        if gv(r, "成本-砖块"): cost["brick"] = int(gv(r, "成本-砖块"))
        if gv(r, "成本-钢材"): cost["steel"] = int(gv(r, "成本-钢材"))
        if gv(r, "成本-窗户"): cost["windows"] = int(gv(r, "成本-窗户"))
        if gv(r, "成本-混凝土"): cost["concrete"] = int(gv(r, "成本-混凝土"))
        maint = int(gv(r, "维护费/分钟") or 0)
        cap = int(gv(r, "住宅容量") or 0)
        cycle = int(gv(r, "生产周期(秒)") or 0)
        out_g = (str(gv(r, "产出-商品") or "").strip())
        out_q = int(gv(r, "产出-数量") or 0)
        inp_g = (str(gv(r, "消耗-商品") or "").strip())
        inp_q_raw = gv(r, "消耗-数量")
        inp_q = inp_q_raw if inp_q_raw is not None else 0  # 可能是 "1,1"(多输入)
        workforce = int(gv(r, "劳动力需求") or 0)
        terrain_raw = (str(gv(r, "特殊地形") or "").strip())
        svc = (str(gv(r, "服务类型") or "").strip())
        radius = int(gv(r, "服务半径(路距离格)") or 0)
        upgrade_to = (str(gv(r, "升级目标建筑ID") or "").strip())
        upgrade_wood = int(gv(r, "升级消耗-木材") or 0)
        upgrade_brick = int(gv(r, "升级消耗-砖块") or 0)
        upgrade_steel = int(gv(r, "升级消耗-钢材") or 0)
        upgrade_windows = int(gv(r, "升级消耗-窗户") or 0)
        upgrade_concrete = int(gv(r, "升级消耗-混凝土") or 0)
        extra_raw = gv(r, "额外参数(JSON)")
        extra = json.loads(extra_raw) if extra_raw else {}
        if upgrade_to:
            extra.setdefault("upgrade", {})["to"] = upgrade_to
            extra.setdefault("upgrade", {}).setdefault("cost", {})["wood"] = upgrade_wood
            if upgrade_brick:
                extra.setdefault("upgrade", {}).setdefault("cost", {})["brick"] = upgrade_brick
            if upgrade_steel:
                extra.setdefault("upgrade", {}).setdefault("cost", {})["steel"] = upgrade_steel
            if upgrade_windows:
                extra.setdefault("upgrade", {}).setdefault("cost", {})["windows"] = upgrade_windows
            if upgrade_concrete:
                extra.setdefault("upgrade", {}).setdefault("cost", {})["concrete"] = upgrade_concrete

        # 单资源地形(clay/iron/copper/gold/coast)为锚点匹配字符串;多地形用逗号分隔数组;空=平地
        if terrain_raw in ("coast", "clay", "iron", "copper", "gold", "coal", "zinc", "limestone"):
            terrain_val = terrain_raw
        elif terrain_raw:
            terrain_val = [t.strip() for t in terrain_raw.split(",")]
        else:
            terrain_val = ["plain"]

        d = {
            "id": bid, "name": name, "category": sheet_name.rstrip("类"),
            "tier": tier,
            "cost": cost,
            "terrain": terrain_val,
            "size": {"w": w, "h": h},
        }
        if maint:
            d["maintenance"] = maint
        if cap:
            d["capacity"] = cap
        if cycle:
            prod = {"cycle": cycle, "inputs": {}, "outputs": {}, "workforce": {}}
            if inp_g:
                # [顺序4] 多输入支持:消耗-商品 "a,b" 消耗-数量 "1,1"
                in_goods = [g.strip() for g in inp_g.split(",") if g.strip()]
                in_qtys = [int(q) for q in str(inp_q).split(",")] if "," in str(inp_q) else [inp_q] * len(in_goods)
                for gi, g in enumerate(in_goods):
                    prod["inputs"][g] = in_qtys[gi] if gi < len(in_qtys) else 1
            if out_g:
                prod["outputs"][out_g] = out_q
            if workforce:
                prod["workforce"][tier or "farmers"] = workforce
            if "radius" in extra:
                prod["radius"] = extra["radius"]
            dev_radius = gv(r, "未开发区域半径")
            if dev_radius:
                prod["radius"] = int(dev_radius)
            d["production"] = prod
        elif workforce:
            d["production"] = {"cycle": 0, "inputs": {}, "outputs": {}, "workforce": {tier or "farmers": workforce}}
        if svc:
            d["service"] = {"type": svc, "radius": radius}
            if svc == "warehouse": d["special"] = "warehouse"
        if "special" in extra:
            d["special"] = extra["special"]
        if "upgrade" in extra:
            d["upgrade"] = extra["upgrade"]
        BUILDINGS[bid] = d

body = json.dumps(BUILDINGS, ensure_ascii=False, indent=2)
js = """/* buildings-data.js — 建筑参数数据表【由 tools/gen-buildings-js.py 从 data/buildings-config.xlsx 生成】
 * 手动改此文件无效——请编辑 data/buildings-config.xlsx 后重新生成。
 */
(function (root) {
  'use strict';
  const BUILDINGS = %s;
  const api = { BUILDINGS };
  root.Engine = root.Engine || {};
  root.Engine.buildingsData = api;
  if (typeof module !== 'undefined' && module.exports) module.exports = api;
})(typeof globalThis !== 'undefined' ? globalThis : this);
""" % body
open(OUT, "w", encoding="utf-8").write(js)
print("OK: %s (%d 个建筑)" % (OUT, len(BUILDINGS)))
