# -*- coding: utf-8 -*-
"""B-63:向 data/buildings-config.xlsx 追加海事建筑(制帆厂/帆船造船厂/码头)
数值来源:multi-island-decision-ledger.md MI-08~MI-10(用户批准)"""
import openpyxl

XLSX = "data/buildings-config.xlsx"
wb = openpyxl.load_workbook(XLSX)
ws = wb["生产类"]

rows = [
    # id, 名称, 阶层, 宽, 高, 金币, 木, 砖, 钢, 窗, 混, 维护, 劳动力, 产出, 周期, 产出数, 消耗, 消耗数, 特殊地形, 未开发半径, 服务类型, 服务半径, 备注
    ["sailMaker", "制帆厂", "workers", 5, 5, 500, 8, 10, 0, 0, 0, 75, 50, "sail", 30, 1, "wool", 1, "", "", "", "", "MI-08:工人5×5,500金/8木/10砖,75/min,50工人,羊毛1→船帆1/30秒,普通平地"],
    ["sailingShipyard", "帆船造船厂", "workers", 6, 17, 10000, 20, 25, 0, 0, 0, 100, 100, "", 0, "", "", "", "coast", "", "", "", "MI-09:工人6×17,10000金/20木/25砖,100/min,100工人,海岸,主动订单每单1艘(B-63 ships-data)"],
    ["port", "码头", "workers", 7, 11, 2500, 10, 0, 8, 0, 0, 0, 0, "", 0, "", "", "", "", "", "", "", "MI-10:码头7×11,2500金/10木/8钢,0维护/0劳动力,每岛最多1座,只提供出发权限(B-63 ships-data)"],
]

start = ws.max_row + 1
for i, r in enumerate(rows):
    for c, v in enumerate(r):
        ws.cell(row=start + i, column=c + 1, value=v)

wb.save(XLSX)
print("OK: %s 追加 %d 行(生产类 %d 行)" % (XLSX, len(rows), ws.max_row - 1))
